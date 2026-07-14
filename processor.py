import pandas as pd
import numpy as np
import re
from datetime import datetime
import io
import os
import zipfile


# ─────────────────────────────────────────────
# ZIP Upload Support
# ─────────────────────────────────────────────
#
# Every file uploader in the app (SC Report, Ecom Tracker, Content File,
# AM Exclusion Sheet) also accepts a .zip archive. When a .zip is uploaded,
# every supported data file (.xlsx / .xls / .csv) inside it is extracted and
# loaded individually with the normal loader for that slot, then all of the
# resulting DataFrames are consolidated (concatenated) into a single
# DataFrame before continuing — so a zip full of multiple SC Report exports,
# multiple Ecom Trackers, etc. behaves exactly as if it were one combined file.

_ZIP_SUPPORTED_EXTS = {".xlsx", ".xls", ".csv"}


class _NamedBytesIO(io.BytesIO):
    """A BytesIO buffer that carries a `.name` attribute, mirroring the
    interface loaders expect from Streamlit's UploadedFile objects."""
    def __init__(self, data, name):
        super().__init__(data)
        self.name = name


def is_zip_upload(uploaded_file):
    """True if the uploaded file is a .zip archive (by filename extension)."""
    if uploaded_file is None:
        return False
    return str(uploaded_file.name).lower().endswith(".zip")


def _extract_zip_data_members(uploaded_file):
    """
    Extract every supported data file (.xlsx/.xls/.csv) from a .zip upload.
    Skips folders, hidden files, __MACOSX metadata, and Excel lock files.
    Returns a list of _NamedBytesIO objects, sorted by filename for
    deterministic consolidation order.
    """
    uploaded_file.seek(0)
    members = []
    try:
        with zipfile.ZipFile(uploaded_file) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                base_name = os.path.basename(info.filename)
                if not base_name or base_name.startswith(".") or base_name.startswith("~$"):
                    continue
                if "__MACOSX" in info.filename:
                    continue
                ext = os.path.splitext(base_name)[1].lower()
                if ext not in _ZIP_SUPPORTED_EXTS:
                    continue
                content = zf.read(info)
                members.append(_NamedBytesIO(content, base_name))
    except zipfile.BadZipFile:
        raise ValueError(f"'{uploaded_file.name}' could not be read as a ZIP file — it may be corrupted.")

    members.sort(key=lambda m: m.name.lower())
    return members


def load_consolidated(uploaded_file, loader_fn, **loader_kwargs):
    """
    Load `uploaded_file` with `loader_fn`. If `uploaded_file` is a .zip
    archive, every supported data file inside is extracted, individually
    loaded with `loader_fn(member, **loader_kwargs)`, and the resulting
    DataFrames are concatenated (consolidated) into one DataFrame.

    Any error while loading a specific member is re-raised with the member's
    filename (and the parent zip's filename) attached for clarity.
    """
    if uploaded_file is None:
        return None

    if not is_zip_upload(uploaded_file):
        return loader_fn(uploaded_file, **loader_kwargs)

    members = _extract_zip_data_members(uploaded_file)
    if not members:
        raise ValueError(
            f"ZIP file '{uploaded_file.name}' does not contain any supported data files "
            f"(.xlsx, .xls, .csv)."
        )

    frames = []
    for member in members:
        try:
            df = loader_fn(member, **loader_kwargs)
        except EcomTrackerSheetNotFoundError as e:
            raise EcomTrackerSheetNotFoundError(
                f"Inside ZIP '{uploaded_file.name}', file '{member.name}': {e}"
            )
        except Exception as e:
            raise ValueError(
                f"Failed to read '{member.name}' inside ZIP '{uploaded_file.name}': {e}"
            )
        frames.append(df)

    consolidated = pd.concat(frames, ignore_index=True, sort=False)
    return consolidated


# ─────────────────────────────────────────────
# File Loading
# ─────────────────────────────────────────────

def load_file(uploaded_file, sheet_name=0, header=0, dtype=None):
    """Load an uploaded Streamlit file (xlsx/xls/csv) into a DataFrame."""
    name = uploaded_file.name.lower()
    uploaded_file.seek(0)
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file, dtype=dtype)
    else:
        return pd.read_excel(uploaded_file, sheet_name=sheet_name, header=header, dtype=dtype)


def _detect_ecom_header_row(raw_df, max_scan=10):
    """
    Scan the first max_scan rows of a raw (header=None) DataFrame to find
    the row containing the Ecom Tracker header (identified by STYLE# or common
    marker columns). Returns the 0-based row index, or 3 as fallback.
    """
    markers = {"style#", "style #", "style no", "sku", "article"}
    for i in range(min(max_scan, len(raw_df))):
        row_vals = [str(v).strip().lower() for v in raw_df.iloc[i] if pd.notna(v)]
        if any(m in row_vals for m in markers):
            return i
    return 3  # documented default: headers on row 4 (0-indexed: 3)


def _load_sheet_raw(uploaded_file, sheet_name):
    """Load a sheet with header=None to allow manual header detection."""
    uploaded_file.seek(0)
    return pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)


class EcomTrackerSheetNotFoundError(ValueError):
    """Raised when the region-specific worksheet is missing from the Ecom Tracker."""
    pass


def load_ecom_tracker(uploaded_file, region):
    """
    Load Ecom Tracker, auto-detecting the true header row by scanning for STYLE#.
    Falls back to row 4 (index 3) per spec if not found.

    The workbook must contain a worksheet whose name exactly matches the selected
    Region (case-insensitive), e.g. Region = "PH" -> reads the "PH" tab. Every
    downstream field, lookup, mapping, and validation is then sourced exclusively
    from that worksheet. If the required region tab is missing, a clear
    EcomTrackerSheetNotFoundError is raised rather than silently falling back to
    a different worksheet.
    """
    uploaded_file.seek(0)
    name = uploaded_file.name.lower()

    # NOTE: pandas' default NA-value list includes "NA", "N/A", etc., which would
    # otherwise silently convert those Launch Date cells to real NaN before our
    # code ever sees them. keep_default_na=False (with an empty na_values list)
    # preserves them as literal text so the "NA = eligible" rule can be applied.
    if name.endswith(".csv"):
        # CSV files have no worksheets/tabs, so region-tab validation doesn't apply.
        raw = pd.read_csv(uploaded_file, header=None, keep_default_na=False, na_values=[])
        header_row = _detect_ecom_header_row(raw)
        uploaded_file.seek(0)
        return pd.read_csv(uploaded_file, header=header_row, keep_default_na=False, na_values=[])

    # ── Resolve the region-specific worksheet (case-insensitive exact match) ──
    try:
        xls = pd.ExcelFile(uploaded_file)
        sheet_names = xls.sheet_names
    except Exception as e:
        raise ValueError(f"Could not read the Ecom Tracker workbook: {e}")

    sheet_lookup = {str(s).strip().lower(): s for s in sheet_names}
    resolved_sheet = sheet_lookup.get(str(region).strip().lower())

    if resolved_sheet is None:
        raise EcomTrackerSheetNotFoundError(
            f"The required '{region}' tab was not found in the uploaded Ecom Tracker file. "
            f"Available tabs: {', '.join(sheet_names)}. "
            f"Please upload an Ecom Tracker that contains a worksheet named exactly '{region}' "
            f"for the selected region."
        )

    uploaded_file.seek(0)
    raw = _load_sheet_raw(uploaded_file, resolved_sheet)
    header_row = _detect_ecom_header_row(raw)
    uploaded_file.seek(0)
    df = pd.read_excel(
        uploaded_file, sheet_name=resolved_sheet, header=header_row,
        keep_default_na=False, na_values=[],
    )
    return df


def load_am_exclusion(uploaded_file, region):
    """Load AM Exclusion sheet based on region."""
    sheet_map = {"SG": "SG VC Exclusions", "MY": "MY VC Exclusions", "PH": "PH VC Exclusions"}
    sheet_name = sheet_map.get(region, 0)
    uploaded_file.seek(0)
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    try:
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
    except Exception:
        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, sheet_name=0)
    return df


# ─────────────────────────────────────────────
# Column Utilities
# ─────────────────────────────────────────────

def col_letter_to_index(letter):
    """Convert Excel column letter(s) to 0-based index. e.g. 'A'->0, 'AX'->49"""
    letter = letter.upper()
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1


def find_col(df, candidates):
    """Return the first column name from df that matches any candidate (case-insensitive)."""
    cols_lower = {c.strip().lower(): c for c in df.columns}
    for cand in candidates:
        key = cand.strip().lower()
        if key in cols_lower:
            return cols_lower[key]
    return None


def normalize_str(s):
    if pd.isna(s):
        return ""
    return str(s).strip()


# ─────────────────────────────────────────────
# Price Tier Column Parsing
# ─────────────────────────────────────────────

def _find_tier_col_by_name(tier_df_cols, candidates):
    """Case-insensitive name search within a list of column names."""
    lower_map = {c.strip().lower(): c for c in tier_df_cols}
    for cand in candidates:
        key = cand.strip().lower()
        if key in lower_map:
            return lower_map[key]
    return None


def parse_price_tier_ref(ref, ecom_df):
    """
    Resolve price tier columns from a reference like 'AX-BA'.

    Two-pass strategy:
      Pass 1 — name-based: search for RRP/SRP/DISC%/Exclusion header names.
      Pass 2 — positional fallback using absolute or relative indices.

    Fixed order within the range:
        START   = RRP
        START+1 = SRP
        START+2 = DISC %
        END     = Exclusion (always last column in range)
    """
    ref = ref.strip().upper()
    if "-" not in ref:
        raise ValueError(f"Invalid Price Tier Reference: '{ref}'. Expected format like 'AX-BA'.")

    parts = ref.split("-")
    start_letter = parts[0].strip()
    end_letter   = parts[1].strip()

    start_abs = col_letter_to_index(start_letter)
    end_abs   = col_letter_to_index(end_letter)

    cols   = list(ecom_df.columns)
    n_cols = len(cols)

    # Pass 1: name-based
    rrp_col       = _find_tier_col_by_name(cols, ["RRP", "Retail Price", "Retail Selling Price"])
    srp_col       = _find_tier_col_by_name(cols, ["SRP", "Selling Price", "Sale Price"])
    disc_col      = _find_tier_col_by_name(cols, ["DISC %", "DISC%", "Discount %", "Discount%", "Disc %"])
    exclusion_col = _find_tier_col_by_name(cols, ["Exclusion", "Exclusion Remarks", "Excl", "VC Exclusion", "Voucher Exclusion"])

    # Pass 2: positional fallback
    span = end_abs - start_abs + 1
    if not all([rrp_col, srp_col, disc_col, exclusion_col]):
        # Strategy a: absolute indices fit
        if start_abs < n_cols and end_abs < n_cols:
            tier_cols_pos = cols[start_abs: end_abs + 1]
            if len(tier_cols_pos) >= 4:
                if not rrp_col:       rrp_col       = tier_cols_pos[0]
                if not srp_col:       srp_col       = tier_cols_pos[1]
                if not disc_col:      disc_col      = tier_cols_pos[2]
                if not exclusion_col: exclusion_col = tier_cols_pos[-1]

        # Strategy b: last N columns
        if not all([rrp_col, srp_col, disc_col, exclusion_col]):
            tail = cols[max(0, n_cols - span):]
            if len(tail) >= 4:
                if not rrp_col:       rrp_col       = tail[0]
                if not srp_col:       srp_col       = tail[1]
                if not disc_col:      disc_col      = tail[2]
                if not exclusion_col: exclusion_col = tail[-1]

        # Strategy c: last 4 cols
        if not all([rrp_col, srp_col, disc_col, exclusion_col]):
            if n_cols >= 4:
                offset = max(0, n_cols - 4)
                if not rrp_col:       rrp_col       = cols[offset]
                if not srp_col:       srp_col       = cols[offset + 1]
                if not disc_col:      disc_col      = cols[offset + 2]
                if not exclusion_col: exclusion_col = cols[-1]
            else:
                raise ValueError(
                    f"Ecom Tracker has only {n_cols} column(s) — cannot extract 4 price tier columns. "
                    f"Please check the correct sheet/file is uploaded and headers start at row 4."
                )

    try:
        s = cols.index(rrp_col)
        e = cols.index(exclusion_col)
        tier_cols = cols[s: e + 1]
    except ValueError:
        tier_cols = [rrp_col, srp_col, disc_col, exclusion_col]

    return {
        "tier_cols":      tier_cols,
        "rrp_col":        rrp_col,
        "srp_col":        srp_col,
        "disc_col":       disc_col,
        "exclusion_col":  exclusion_col,
        "start_idx":      start_abs,
        "end_idx":        end_abs,
        "start_letter":   start_letter,
        "end_letter":     end_letter,
    }


# ─────────────────────────────────────────────
# Date Processing
# ─────────────────────────────────────────────
#
# The Ecom Tracker stores Launch Dates in DD-MM-YYYY format.
# When pandas reads an Excel date column it converts cells to Python datetime
# objects, so we handle that case first before falling back to string parsing.

INVALID_DATE_TEXTS = {"past season", "tbc", "00-jan-1900", ""}

def parse_launch_date(val):
    """
    Return a datetime.date or None.

    Priority order:
      1. Already a date/datetime object (pandas Excel read) → use directly.
      2. String in DD-MM-YYYY format (primary Ecom Tracker format) → parse first.
      3. Other common date strings → try remaining formats.
      4. Reject: blank, text-only, 00-Jan-1900, 'Past Season', 'TBC', etc.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None

    # ── Path 1: already a Python date / datetime (pandas Excel conversion) ──
    if isinstance(val, datetime):
        # Guard against Excel's epoch ghost date 1900-01-00
        if val.year == 1900 and val.month == 1 and val.day <= 1:
            return None
        return val.date()
    if hasattr(val, 'date') and callable(val.date):
        # covers pandas Timestamp and datetime subclasses
        try:
            d = val.date()
            if d.year == 1900 and d.month == 1 and d.day <= 1:
                return None
            return d
        except Exception:
            pass

    # ── Path 2: string parsing ──
    try:
        is_na = pd.isna(val)
    except Exception:
        is_na = False
    if is_na:
        return None

    s = str(val).strip()
    if not s or s.lower() in INVALID_DATE_TEXTS:
        return None
    # Reject pure-text values (no digit = not a date)
    if not any(ch.isdigit() for ch in s):
        return None
    # Reject Excel ghost date variants
    if "1900" in s and ("jan" in s.lower() or s.startswith("00")):
        return None

    # Try DD-MM-YYYY first (primary Ecom Tracker format), then others
    for fmt in (
        "%d-%m-%Y",   # ← PRIMARY: DD-MM-YYYY (Ecom Tracker)
        "%d/%m/%Y",   # DD/MM/YYYY
        "%Y-%m-%d",   # ISO (pandas default str representation)
        "%d %b %Y",   # 13 May 2025
        "%d-%b-%Y",   # 13-May-2025
        "%d/%b/%Y",   # 13/May/2025
        "%m/%d/%Y",   # US format fallback
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass

    # Last resort: let pandas infer (explicit formats above already cover DD-MM-YYYY)
    try:
        return pd.to_datetime(s, dayfirst=False).date()
    except Exception:
        pass

    return None


def format_launch_date(val):
    """
    Return the launch date as a DD-MM-YYYY string, or empty string if invalid.
    Preserves the Ecom Tracker's native DD-MM-YYYY format throughout.
    """
    d = parse_launch_date(val)
    if d is None:
        return ""
    return d.strftime("%d-%m-%Y")


_NA_TEXTS = {"na", "n/a", "n.a", "n.a."}


def format_launch_date_or_na(val):
    """
    Same as format_launch_date, but preserves an explicit "NA" value instead of
    blanking it out. "NA" Launch Dates are treated as eligible during Launch
    Date validation (see evaluate_sku_for_voucher, Filter 1).
    """
    try:
        is_na_val = pd.isna(val)
    except Exception:
        is_na_val = False
    if not is_na_val:
        s = str(val).strip().lower()
        if s in _NA_TEXTS:
            return "NA"
    return format_launch_date(val)


# ─────────────────────────────────────────────
# Region Configuration — RRP / SRP thresholds
# ─────────────────────────────────────────────
#
# RRP: exclude when RRP <= threshold, RRP == 0, or RRP is blank. Eligible when RRP > threshold.
# SRP: eligible when SRP == 0 or SRP > threshold. Exclude when 0 < SRP <= threshold.
REGION_PRICE_THRESHOLDS = {
    "SG": 16.0,
    "MY": 38.9,
    "PH": 649.0,
}


# ─────────────────────────────────────────────
# Voucher Percentage Parsing
# ─────────────────────────────────────────────

def extract_voucher_pct(name):
    """
    Extract the leading numeric percentage from a voucher name.
    Supports: '10% VC' -> 10.0, '30 % NMS' -> 30.0, '40 percent' -> 40.0
    Returns None if no percentage found.
    """
    m = re.search(r"(\d+(?:\.\d+)?)\s*(?:%|percent)", name.strip(), re.IGNORECASE)
    if m:
        return float(m.group(1))
    return None


# ─────────────────────────────────────────────
# AM Exclusion Logic  (fully rewritten)
# ─────────────────────────────────────────────
#
# Rule 1 — Global exclusion keywords → exclude from ALL vouchers
# Rule 2 — "Exclude from X%" → exclude from vouchers with exactly X%
# Rule 3 — "Exclude from X% and above" → exclude from vouchers with % >= X
#
# Multiple rules can apply to one ALU; all are applied (most restrictive wins).

# Global exclusion keyword fragments (case-insensitive substring match)
_GLOBAL_EXCL_KEYWORDS = [
    "exclude from all voucher",
    "exclude from platform voucher",
    "exclude from vc",
    "exclude from voucher",
    "voucher exclusion",
]

# Bare "Exclude" or "Exclude from VC" with no percentage
_BARE_EXCLUDE_RE = re.compile(
    r"^\s*exclude\s*$"
    r"|exclude\s+from\s+vc\s*$"
    r"|exclude\s+from\s+voucher\s*$",
    re.IGNORECASE,
)

# "Exclude from X% and above" / "Exclude from X% above"
_EXCL_GTE_RE = re.compile(
    r"exclude.*?(\d+(?:\.\d+)?)\s*(?:%|percent).*?\b(and\s+)?above\b",
    re.IGNORECASE,
)

# "Exclude from X%" (specific, no "above") — catches "Exclude from 10% VC",
# "Exclude from 15%", "Exclude from 20% Voucher" etc.
_EXCL_EXACT_RE = re.compile(
    r"exclude.*?(\d+(?:\.\d+)?)\s*(?:%|percent)",
    re.IGNORECASE,
)


def parse_am_exclusion_rules(excl_text):
    """
    Parse an AM Exclusion cell value and return a list of rule dicts.
    Each rule has the form: {"mode": "all"|"exact"|"gte", "threshold": float|None}

    Multiple rules may apply to a single ALU — caller applies them all.
    """
    if pd.isna(excl_text) or not str(excl_text).strip():
        return []

    text = str(excl_text).strip()
    rules = []

    # ── Rule 1: global exclusion (keyword match) ──
    text_lower = text.lower()
    is_global = any(kw in text_lower for kw in _GLOBAL_EXCL_KEYWORDS)
    if not is_global and _BARE_EXCLUDE_RE.search(text):
        is_global = True
    if is_global:
        return [{"mode": "all", "threshold": None}]

    # ── Rule 3: "X% and above" ── (must check before exact to avoid false match)
    for m in _EXCL_GTE_RE.finditer(text):
        rules.append({"mode": "gte", "threshold": float(m.group(1))})

    # ── Rule 2: exact percentage (lines that do NOT also contain "above") ──
    # Strip already-matched "and above" segments before scanning for exact
    text_no_above = _EXCL_GTE_RE.sub("", text)
    for m in _EXCL_EXACT_RE.finditer(text_no_above):
        pct = float(m.group(1))
        # Avoid duplicating a threshold already captured by Rule 3
        already_gte = any(r["threshold"] == pct and r["mode"] == "gte" for r in rules)
        if not already_gte:
            rules.append({"mode": "exact", "threshold": pct})

    return rules


def is_am_excluded_for_voucher(rules, voucher_pct):
    """
    Given a list of parsed AM exclusion rules and the voucher's numeric %,
    return True if the ALU should be excluded from that voucher.

    If voucher_pct is None (no % in voucher name), only global rules apply.
    """
    for rule in rules:
        mode      = rule["mode"]
        threshold = rule["threshold"]

        if mode == "all":
            return True
        if mode == "exact" and voucher_pct is not None and voucher_pct == threshold:
            return True
        if mode == "gte" and voucher_pct is not None and voucher_pct >= threshold:
            return True

    return False


def am_exclusion_label(rules):
    """Return a human-readable label for the AM Exclude column from a rule list."""
    if not rules:
        return ""
    if any(r["mode"] == "all" for r in rules):
        return "Exclude from all Vouchers"
    parts = []
    for r in rules:
        if r["mode"] == "exact":
            t = int(r["threshold"]) if r["threshold"] == int(r["threshold"]) else r["threshold"]
            parts.append(f"Exclude from {t}%")
        elif r["mode"] == "gte":
            t = int(r["threshold"]) if r["threshold"] == int(r["threshold"]) else r["threshold"]
            parts.append(f"Exclude from {t}% and above")
    return " | ".join(parts) if parts else ""


# ─────────────────────────────────────────────
# Exclusion Remarks Extraction (for dropdown)
# ─────────────────────────────────────────────

def get_exclusion_remarks(ecom_df, exclusion_col):
    """
    Extract unique, non-blank values from the Exclusion column of the Ecom Tracker.
    Used to populate the inclusion keyword dropdown in the UI.

    Deduplication is case-insensitive: if the Ecom Tracker contains both
    'OPEN FOR ALL' and 'Open for all', only one representative value is kept
    (the first one encountered, uppercased for consistency). This mirrors the
    case-insensitive matching used in evaluate_sku_for_voucher (Filter 5).

    Returns a sorted list of strings.
    """
    if not exclusion_col or exclusion_col not in ecom_df.columns:
        return []
    raw_vals = (
        ecom_df[exclusion_col]
        .dropna()
        .astype(str)
        .str.strip()
        .replace("", pd.NA)
        .dropna()
        .tolist()
    )
    # Deduplicate case-insensitively, keeping the uppercase form as canonical
    seen_lower = {}
    for v in raw_vals:
        key = v.lower()
        if key not in seen_lower:
            seen_lower[key] = v.upper()  # store uppercase as canonical display value
    return sorted(seen_lower.values())


# ─────────────────────────────────────────────
# Core Processing Function
# ─────────────────────────────────────────────

def _prepare_common_data(
    marketplace,
    region,
    sc_report_file,
    ecom_file,
    content_file,
    am_excl_file,
    price_tier_ref,
    cutoff_date,
    launch_date_col_override=None,
):
    """
    Shared loading / mapping / base-column pipeline used by every marketplace
    (Lazada, Zalora, Shopee, ...). Returns a dict with everything downstream
    processing needs:

      out              - SC Report DataFrame with ALU + base working columns added
                          (Launch Date, Ecom Status, RRP, SRP, DISC %, Exclusion, AM Exclude)
      warnings         - list of info/warning strings
      ecom_df          - the loaded Ecom Tracker (region worksheet only)
      ecom_lookup_dict - ALU -> {column: first non-blank value}
      get_ecom_val     - fn(alu, col) -> value
      am_excl_map      - ALU -> [rule dicts]
      price_threshold  - region RRP/SRP threshold
      tier_info        - resolved price tier column mapping
      style_col, launch_date_col, ecom_status_col - resolved Ecom Tracker columns
    """
    warnings = []

    # ── 1. Load SC Report (zip archives are extracted & consolidated) ──
    sc_df = load_consolidated(sc_report_file, load_file, dtype=str)
    sc_df.columns = sc_df.columns.str.strip()
    if is_zip_upload(sc_report_file):
        warnings.append(
            f"ℹ️ SC Report: ZIP archive '{sc_report_file.name}' extracted and consolidated "
            f"into {len(sc_df):,} row(s)."
        )

    # ── 2. Load Content File (zip archives are extracted & consolidated) ──
    content_df = load_consolidated(content_file, load_file, dtype=str)
    content_df.columns = content_df.columns.str.strip()
    if is_zip_upload(content_file):
        warnings.append(
            f"ℹ️ Content File: ZIP archive '{content_file.name}' extracted and consolidated "
            f"into {len(content_df):,} row(s)."
        )

    # ── 3. Load Ecom Tracker (zip archives are extracted & consolidated) ──
    ecom_df = load_consolidated(ecom_file, load_ecom_tracker, region=region)
    ecom_df.columns = [str(c).strip() for c in ecom_df.columns]
    if is_zip_upload(ecom_file):
        warnings.append(
            f"ℹ️ Ecom Tracker: ZIP archive '{ecom_file.name}' extracted and consolidated "
            f"into {len(ecom_df):,} row(s) (region '{region}' tab from each file)."
        )

    # ── 4. Load AM Exclusion (optional; zip archives are extracted & consolidated) ──
    am_df = None
    if am_excl_file is not None:
        am_df = load_consolidated(am_excl_file, load_am_exclusion, region=region)
        am_df.columns = am_df.columns.str.strip()
        if is_zip_upload(am_excl_file):
            warnings.append(
                f"ℹ️ AM Exclusion Sheet: ZIP archive '{am_excl_file.name}' extracted and "
                f"consolidated into {len(am_df):,} row(s)."
            )

    # ── 5. Parse Price Tier Reference ──
    tier_info     = parse_price_tier_ref(price_tier_ref, ecom_df)
    rrp_col       = tier_info["rrp_col"]
    srp_col       = tier_info["srp_col"]
    disc_col      = tier_info["disc_col"]
    exclusion_col = tier_info["exclusion_col"]

    # Info message showing resolved column mapping
    cols_list = list(ecom_df.columns)
    def _col_label(col_name):
        try:
            idx = cols_list.index(col_name)
            n, letters = idx + 1, ""
            while n:
                n, r = divmod(n - 1, 26)
                letters = chr(65 + r) + letters
            return f"{letters} ({col_name})"
        except ValueError:
            return col_name

    warnings.append(
        f"ℹ️ Price Tier column mapping — "
        f"RRP: {_col_label(rrp_col)} | "
        f"SRP: {_col_label(srp_col)} | "
        f"DISC %: {_col_label(disc_col)} | "
        f"Exclusion: {_col_label(exclusion_col)}"
    )

    # ── 6. Identify key columns in Content File ──
    ean_col         = find_col(content_df, ["EAN", "EAN Code", "EAN Number"])
    alu_col_content = find_col(content_df, ["Color No", "Colour No", "Color No.", "Color Number", "ALU"])
    if not ean_col:
        raise ValueError("Content File: Cannot find 'EAN' column.")
    if not alu_col_content:
        raise ValueError("Content File: Cannot find 'Color No' / ALU column.")

    content_map = (
        content_df[[ean_col, alu_col_content]]
        .dropna(subset=[ean_col])
        .drop_duplicates(subset=[ean_col])
        .set_index(ean_col)[alu_col_content]
        .to_dict()
    )

    # ── 7. Identify SellerSKU column in SC Report ──
    seller_sku_col = find_col(sc_df, ["SellerSKU", "Seller SKU", "SKU", "seller_sku"])
    if not seller_sku_col:
        raise ValueError("SC Report: Cannot find 'SellerSKU' column.")

    # ── 8. Map ALU onto SC Report ──
    sc_df["ALU"] = sc_df[seller_sku_col].map(content_map).fillna("")

    # ── 9. Identify key columns in Ecom Tracker ──
    style_col = find_col(ecom_df, ["STYLE#", "Style#", "STYLE #", "Style No", "StyleNo", "ALU"])
    if not style_col:
        raise ValueError("Ecom Tracker: Cannot find 'STYLE#' column.")

    if launch_date_col_override and launch_date_col_override in ecom_df.columns:
        launch_date_col = launch_date_col_override
    else:
        launch_date_col = find_col(
            ecom_df,
            ["Launch Dates", "Launch Date", "LaunchDate", "Launch date",
             "LAUNCH DATE", "LAUNCH DATES", "Ecom Launch Date"],
        )
        if launch_date_col_override and launch_date_col_override not in ecom_df.columns:
            warnings.append(
                f"Ecom Tracker: selected Launch Date column '{launch_date_col_override}' not found — "
                f"falling back to auto-detected column '{launch_date_col or 'none'}'."
            )
    if not launch_date_col:
        warnings.append("Ecom Tracker: 'Launch Dates' column not found — Launch Date will be blank.")

    ecom_status_col = find_col(
        ecom_df,
        [marketplace, marketplace.upper(), f"{marketplace} Status", f"{marketplace} Ecom Status"],
    )

    if not ecom_status_col:
        warnings.append(
            f"Ecom Tracker: '{marketplace}' column not found — Ecom Status will be blank. "
            f"Expected a column named '{marketplace}'."
        )

    # ── 10. Build ALU → Ecom Tracker row lookup ──
    # Uses first-non-blank value per (ALU, column) across ALL rows sharing that ALU.
    # drop_duplicates() would discard duplicate ALU rows and potentially pick a row
    # with blank Ecom Status or a future launch date — this approach avoids that.
    ecom_df[style_col] = ecom_df[style_col].astype(str).str.strip()

    def _build_ecom_lookup(df, key_col):
        def _is_blank(v):
            if v is None:
                return True
            try:
                if pd.isna(v):
                    return True
            except Exception:
                pass
            return str(v).strip() == ""

        lookup = {}
        for _, row in df.iterrows():
            alu = str(row[key_col]).strip()
            if not alu or alu == "nan":
                continue
            if alu not in lookup:
                lookup[alu] = {}
            for col_name, val in row.items():
                if col_name == key_col:
                    continue
                # Keep the first non-blank value seen for each column
                if col_name not in lookup[alu] or _is_blank(lookup[alu].get(col_name)):
                    if not _is_blank(val):
                        lookup[alu][col_name] = val
        return lookup

    ecom_lookup_dict = _build_ecom_lookup(ecom_df, style_col)

    def get_ecom_val(alu, col):
        if not col or not alu or alu not in ecom_lookup_dict:
            return ""
        val = ecom_lookup_dict[alu].get(col, "")
        if val is None:
            return ""
        try:
            if pd.isna(val):
                return ""
        except Exception:
            pass
        return val

    # ── 11. AM Exclusion mapping: ALU → list of rule dicts ──
    #
    # Matching is done via ALU. A Marketplace column (if present) is used to
    # scope a row to the currently selected Marketplace, using case-insensitive
    # multi-delimiter matching (e.g. "Lazada, Zalora" / "Lazada & Zalora").
    # Rows with a blank Marketplace value are treated as applicable to all
    # marketplaces.
    #
    # Fallback rule: if an ALU is present in the AM Exclusion Sheet (and
    # applicable to this marketplace) but its Exclusion Type text does not
    # resolve into any recognised criteria, the ALU defaults to being excluded
    # from ALL vouchers.
    am_excl_map = {}   # ALU → [{"mode":..., "threshold":...}, ...]
    if am_df is not None:
        article_col      = find_col(am_df, ["Article", "ALU", "SKU", "Style No"])
        excl_type_col    = find_col(am_df, ["Exclusion Type", "ExclusionType", "Exclusion", "Type"])
        marketplace_col  = find_col(am_df, ["Marketplace", "Platform", "MP", "Channel", "Ecom"])
        _mp_split_re     = re.compile(r"[,/;&]+|\band\b", re.IGNORECASE)

        if article_col:
            for _, row in am_df.iterrows():
                alu_key = normalize_str(row[article_col])
                if not alu_key:
                    continue

                # Marketplace scoping (case-insensitive, multi-delimiter)
                if marketplace_col:
                    mkt_val = normalize_str(row[marketplace_col])
                    if mkt_val:
                        tokens = [
                            t.strip().lower()
                            for t in _mp_split_re.split(mkt_val)
                            if t.strip()
                        ]
                        mp_lower = marketplace.lower()
                        if not any(mp_lower == t or mp_lower in t or t in mp_lower for t in tokens):
                            continue  # row does not apply to the selected marketplace

                excl_text = normalize_str(row[excl_type_col]) if excl_type_col else ""
                rules = parse_am_exclusion_rules(excl_text)
                if not rules:
                    # ALU present in the AM Exclusion Sheet but no configured criteria
                    # matched → default: exclude from ALL vouchers.
                    rules = [{"mode": "all", "threshold": None}]

                if alu_key in am_excl_map:
                    am_excl_map[alu_key].extend(rules)   # merge multiple rows
                else:
                    am_excl_map[alu_key] = rules

    # ── 12. Build output DataFrame ──
    out  = sc_df.copy()
    alus = out["ALU"].astype(str).str.strip()

    # Launch Date — "NA" values are preserved (treated as eligible; see Filter 1)
    out["Launch Date"] = (
        alus.map(lambda a: get_ecom_val(a, launch_date_col)).map(format_launch_date_or_na)
        if launch_date_col else ""
    )

    # Ecom Status
    out["Ecom Status"] = (
        alus.map(lambda a: normalize_str(get_ecom_val(a, ecom_status_col)))
        if ecom_status_col else ""
    )

    # Price fields
    out["RRP"]    = alus.map(lambda a: get_ecom_val(a, rrp_col))    if rrp_col    else ""
    out["SRP"]    = alus.map(lambda a: get_ecom_val(a, srp_col))    if srp_col    else ""
    out["DISC %"] = alus.map(lambda a: get_ecom_val(a, disc_col))   if disc_col   else ""

    # Ecom Tracker Exclusion / Inclusion Remarks
    out["Exclusion"] = alus.map(lambda a: normalize_str(get_ecom_val(a, exclusion_col)))

    # AM Exclude — human-readable label
    out["AM Exclude"] = alus.map(
        lambda a: am_exclusion_label(am_excl_map.get(a, []))
    )

    price_threshold = REGION_PRICE_THRESHOLDS.get(region, 16.0)

    return {
        "out": out,
        "warnings": warnings,
        "sc_df": sc_df,
        "content_df": content_df,
        "ecom_df": ecom_df,
        "ecom_lookup_dict": ecom_lookup_dict,
        "get_ecom_val": get_ecom_val,
        "am_excl_map": am_excl_map,
        "price_threshold": price_threshold,
        "tier_info": tier_info,
        "rrp_col": rrp_col,
        "srp_col": srp_col,
        "disc_col": disc_col,
        "exclusion_col": exclusion_col,
        "style_col": style_col,
        "launch_date_col": launch_date_col,
        "ecom_status_col": ecom_status_col,
        "seller_sku_col": seller_sku_col,
    }


def process_voucher_eligibility(
    marketplace,
    region,
    sc_report_file,
    ecom_file,
    content_file,
    am_excl_file,
    price_tier_ref,
    cutoff_date,
    voucher_configs,
    launch_date_col_override=None,
):
    """
    Main processing function for Lazada / Zalora (ALU-level eligibility).
    Returns (output_df, warnings_list).

    Each voucher in voucher_configs is evaluated completely independently:
      1. Launch Date filter
      2. Ecom Status filter
      3. RRP filter
      4. SRP filter
      5. Inclusion Keyword filter (exact match against Exclusion column values)
      6. AM Exclusion filter (per-voucher, using parsed rules)

    Eligible SKUs are marked "Yes-Eligible"; ineligible SKUs are left blank.

    NOTE: Shopee uses a completely different, Product-ID-level pipeline —
    see process_shopee_voucher_eligibility().
    """
    common = _prepare_common_data(
        marketplace, region, sc_report_file, ecom_file, content_file, am_excl_file,
        price_tier_ref, cutoff_date, launch_date_col_override,
    )
    out             = common["out"]
    warnings        = common["warnings"]
    ecom_lookup_dict = common["ecom_lookup_dict"]
    am_excl_map     = common["am_excl_map"]
    price_threshold = common["price_threshold"]

    # ── 13. Per-voucher independent eligibility evaluation ──
    def _to_float(val):
        try:
            return float(str(val).replace(",", "").strip())
        except Exception:
            return None

    def evaluate_sku_for_voucher(row, keywords, vc_pct):
        """
        Evaluate one SKU row against one voucher's full criteria independently.
        Returns "Yes-Eligible" if all filters pass, "" otherwise.

        Filter order:
          0. ALU in Ecom Tracker — ALU must exist in the tracker. SKUs with a blank ALU
                                   (EAN not mapped via Content File) or an ALU absent from
                                   the Ecom Tracker are excluded immediately.
          1. Launch Date         — valid DD-MM-YYYY date <= cutoff date. "NA" is treated
                                   as eligible for all regions (SG/MY/PH).
          2. Ecom Status         — must be exactly YES
          3. RRP                 — must be > region price threshold (SG:16 | MY:38.9 | PH:649)
          4. SRP                 — blank/0 is allowed; (0, threshold] inclusive is excluded.
                                   A blank SRP means no sale price (treated as 0 = allowed).
          5. Inclusion Keyword   — Exclusion column must match a selected keyword
                                   (case-insensitive: "Open for all" == "OPEN FOR ALL")
          6. AM Exclusion        — ALU must not be excluded for this voucher %

        Note: DISC % is output for reference only — no filter is applied on it.
        """
        alu = str(row.get("ALU", "")).strip()

        # Filter 0: ALU must exist in Ecom Tracker
        # Covers: blank ALU (EAN→ALU mapping failed) AND ALU not found in tracker at all.
        if not alu or alu not in ecom_lookup_dict:
            return ""

        # Filter 1: Launch Date — "NA" is treated as eligible (cutoff check skipped)
        raw_ld = str(row.get("Launch Date", "")).strip()
        if not raw_ld:
            return ""
        if raw_ld.upper() not in ("NA", "N/A"):
            ld = parse_launch_date(raw_ld)
            if ld is None:
                return ""
            if cutoff_date and ld > cutoff_date:
                return ""

        # Filter 2: Ecom Status
        if str(row.get("Ecom Status", "")).strip().upper() != "YES":
            return ""

        # Filter 3: RRP — must be strictly greater than the region price threshold.
        # RRP <= threshold, RRP == 0, or blank RRP are all excluded.
        rrp = _to_float(row.get("RRP", ""))
        if rrp is None or rrp <= price_threshold:
            return ""

        # Filter 4: SRP — 0 (or blank, treated as 0) is allowed; SRP > threshold is allowed.
        # 0 < SRP <= threshold is excluded.
        srp = _to_float(row.get("SRP", ""))
        if srp is None:
            srp = 0.0   # blank SRP → no sale price → allowed
        if srp != 0 and srp <= price_threshold:
            return ""

        # Filter 5: Inclusion Keyword — case-insensitive match.
        # The Ecom Tracker stores the same semantic value in mixed casing
        # (e.g. "Open for all" vs "OPEN FOR ALL"). The manual process treats
        # these as identical, so we normalise both sides to lowercase before comparing.
        ecom_excl_val = normalize_str(row.get("Exclusion", "")).lower()
        if keywords:
            if not any(ecom_excl_val == kw.lower() for kw in keywords):
                return ""

        # Filter 6: AM Exclusion
        if alu:
            rules = am_excl_map.get(alu, [])
            if is_am_excluded_for_voucher(rules, vc_pct):
                return ""

        return "Yes-Eligible"

    # ── 14. Build one result column per voucher ──
    for vc in voucher_configs:
        vc_name  = vc.get("name", "").strip()
        keywords = [kw.strip() for kw in vc.get("keywords", []) if str(kw).strip()]
        vc_pct   = extract_voucher_pct(vc_name)

        row_errors: list = []

        def _safe_evaluate(row, _kw=keywords, _pct=vc_pct, _errs=row_errors):
            try:
                return evaluate_sku_for_voucher(row, _kw, _pct)
            except Exception as _e:
                _errs.append(f"Row {row.name}: {_e}")
                return ""

        out[vc_name] = out.apply(_safe_evaluate, axis=1)

        if row_errors:
            sample = row_errors[:5]
            warnings.append(
                f"⚠️ '{vc_name}': {len(row_errors)} row(s) raised errors during evaluation "
                f"(treated as ineligible). Sample errors: {'; '.join(sample)}"
            )

    return out, warnings


# ─────────────────────────────────────────────
# Shopee — Product ID (PID) Level Voucher Eligibility
# ─────────────────────────────────────────────
#
# Shopee eligibility logic is completely different from Lazada/Zalora: every
# check is performed at the Product ID (PID) level, not the ALU level. A
# single Product ID can have multiple SC Report rows (variations) and can be
# associated with one or more ALUs. Whenever ANY ALU tied to a Product ID
# triggers a condition, EVERY row sharing that Product ID inherits the result.

def find_product_id_col(df):
    """Locate the Product ID column on a Shopee Seller Center report."""
    return find_col(df, [
        "Product ID", "ProductID", "Product Id", "Item ID", "ItemID",
        "Item Id", "Parent SKU", "ParentSKU", "Parent Sku",
    ])


def process_shopee_voucher_eligibility(
    region,
    sc_report_file,
    ecom_file,
    content_file,
    am_excl_file,
    price_tier_ref,
    cutoff_date,
    voucher_configs,
    launch_date_col_override=None,
):
    """
    Shopee-specific processing pipeline (Product-ID level).
    Returns (output_df, warnings_list, pid_flag_cols, dynamic_excl_cols).

    Step 1 - Standard mapping/preprocessing (shared with Lazada/Zalora via
             _prepare_common_data): Seller Center enrichment, ALU mapping,
             Product ID identification.
    Step 2 - Four mandatory PID flag columns are added: Future_PID,
             Ecom NO_PID, Low Price_PID, AM Exclude ALU_PID.
    Step 3 - One dynamic "<Exclusion Remark>_PID" column per unique
             Exclusion Remarks value found in the Ecom Tracker.
    Step 4 - Voucher columns are appended last (added by the caller via the
             returned column list, populated below).
    Step 5 - Each PID flag / dynamic column is populated by propagating the
             ALU-level condition to every SC Report row sharing that Product ID.
    Step 6/7 - Voucher eligibility: a Product ID flagged in ANY of the four
             mandatory columns is always ineligible. Otherwise: "Open for all"
             ignores the dynamic columns; any other Inclusion Keyword requires
             a match in its corresponding dynamic column.
    """
    common = _prepare_common_data(
        "Shopee", region, sc_report_file, ecom_file, content_file, am_excl_file,
        price_tier_ref, cutoff_date, launch_date_col_override,
    )
    out              = common["out"]
    warnings         = common["warnings"]
    ecom_df          = common["ecom_df"]
    am_excl_map      = common["am_excl_map"]
    price_threshold  = common["price_threshold"]
    exclusion_col    = common["exclusion_col"]
    seller_sku_col   = common["seller_sku_col"]

    # ── SKU Length Validation (Shopee-specific) ──
    # Shopee's SellerSKU is the authoritative SKU reference for the entire
    # Shopee filtration process. A valid Shopee SellerSKU is always exactly
    # 13 characters long. Rows whose SellerSKU length is anything other than
    # 13 characters are not valid Shopee SKUs (e.g. stray barcodes, partial
    # values, header/footer artifacts) and must be dropped BEFORE any
    # eligibility or voucher filtration logic runs, so they can never
    # contaminate the Product-ID-level flag propagation below.
    sku_len = out[seller_sku_col].astype(str).str.strip().str.len()
    valid_len_mask = sku_len == 13
    n_dropped = int((~valid_len_mask).sum())
    if n_dropped:
        warnings.append(
            f"ℹ️ SC Report: removed {n_dropped:,} row(s) where '{seller_sku_col}' "
            f"(SellerSKU) length was not exactly 13 characters — only valid "
            f"13-character Shopee SellerSKUs are kept before filtration."
        )
    out = out.loc[valid_len_mask].reset_index(drop=True)

    # ── Identify the Product ID column on the Shopee SC Report ──
    pid_col = find_product_id_col(out)
    if not pid_col:
        raise ValueError(
            "Shopee SC Report: cannot find a 'Product ID' column. "
            "Expected one of: Product ID, Item ID, Parent SKU."
        )
    out["Product ID"] = out[pid_col].astype(str).str.strip()
    alus = out["ALU"].astype(str).str.strip()

    def _to_float(val):
        try:
            return float(str(val).replace(",", "").strip())
        except Exception:
            return None

    def _pids_for_mask(mask):
        """Product IDs of every row where an ALU-level condition is True."""
        return set(out.loc[mask, "Product ID"]) - {""}

    # ── Step 5a: Future_PID — Launch Date after the selected cutoff ──
    def _is_future(row):
        raw_ld = str(row.get("Launch Date", "")).strip()
        if not raw_ld or raw_ld.upper() in ("NA", "N/A"):
            return False
        ld = parse_launch_date(raw_ld)
        if ld is None:
            return False
        return bool(cutoff_date) and ld > cutoff_date

    future_pids = _pids_for_mask(out.apply(_is_future, axis=1))

    # ── Step 5b: Ecom NO_PID — Ecom Status is anything other than "Yes" ──
    ecom_no_mask = out["Ecom Status"].astype(str).str.strip().str.upper() != "YES"
    ecom_no_pids = _pids_for_mask(ecom_no_mask)

    # ── Step 5c: Low Price_PID — fails the existing RRP/SRP validation ──
    def _is_low_price(row):
        rrp = _to_float(row.get("RRP", ""))
        if rrp is None or rrp <= price_threshold:
            return True
        srp = _to_float(row.get("SRP", ""))
        if srp is None:
            srp = 0.0
        return srp != 0 and srp <= price_threshold

    low_price_pids = _pids_for_mask(out.apply(_is_low_price, axis=1))

    # ── Step 5d: AM Exclude ALU_PID — ALU present in the AM Exclusion Sheet ──
    am_excl_mask = alus.map(lambda a: bool(a) and a in am_excl_map)
    am_excl_pids = _pids_for_mask(am_excl_mask)

    out["Future_PID"]         = out["Product ID"].map(lambda p: "FL" if p in future_pids else "")
    out["Ecom NO_PID"]        = out["Product ID"].map(lambda p: "Ecom NO" if p in ecom_no_pids else "")
    out["Low Price_PID"]      = out["Product ID"].map(lambda p: "Low Price" if p in low_price_pids else "")
    out["AM Exclude ALU_PID"] = out["Product ID"].map(lambda p: "AM Exclude" if p in am_excl_pids else "")

    pid_flag_cols = ["Future_PID", "Ecom NO_PID", "Low Price_PID", "AM Exclude ALU_PID"]

    # ── Step 3/5: Dynamic exclusion columns — one per unique Exclusion Remark ──
    unique_remarks = get_exclusion_remarks(ecom_df, exclusion_col) if exclusion_col else []
    excl_series = out["Exclusion"].astype(str).str.strip()

    dynamic_cols = []
    for remark in unique_remarks:
        col_name = f"{remark}_PID"
        remark_pids = _pids_for_mask(excl_series.str.lower() == remark.lower())
        out[col_name] = out["Product ID"].map(lambda p, r=remark, s=remark_pids: r if p in s else "")
        dynamic_cols.append(col_name)

    # ── Step 6/7: Voucher eligibility (PID level) ──
    mandatory_excluded_mask = (
        (out["Future_PID"] != "") |
        (out["Ecom NO_PID"] != "") |
        (out["Low Price_PID"] != "") |
        (out["AM Exclude ALU_PID"] != "")
    )

    for vc in voucher_configs:
        vc_name  = vc.get("name", "").strip()
        keywords = [kw.strip() for kw in vc.get("keywords", []) if str(kw).strip()]

        def _evaluate(row, _kws=keywords):
            # Always exclude PIDs flagged in any of the 4 mandatory columns first.
            if mandatory_excluded_mask.loc[row.name]:
                return ""

            # Case 1 — "Open for all" (or no keyword restriction configured):
            # ignore all dynamic exclusion columns entirely.
            is_open_for_all = (not _kws) or any(
                kw.strip().lower() == "open for all" for kw in _kws
            )
            if is_open_for_all:
                return "Yes-Eligible"

            # Case 2 — specific Inclusion Keyword(s): the PID must be flagged
            # in the matching dynamic "<keyword>_PID" column.
            for kw in _kws:
                col_name = f"{kw}_PID"
                if col_name in out.columns and normalize_str(row.get(col_name, "")).lower() == kw.strip().lower():
                    return "Yes-Eligible"
            return ""

        out[vc_name] = out.apply(_evaluate, axis=1)

    return out, warnings, pid_flag_cols, dynamic_cols


# ─────────────────────────────────────────────
# Excel Output Generation
# ─────────────────────────────────────────────

WORKING_COLS = ["ALU", "Launch Date", "Ecom Status", "RRP", "SRP", "DISC %", "Exclusion", "AM Exclude"]


def generate_excel_output(df, marketplace, voucher_configs):
    """Return bytes of a styled Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"{marketplace} VC Eligibility"

    voucher_names = [vc["name"] for vc in voucher_configs if vc.get("name", "").strip()]

    injected = set(WORKING_COLS + voucher_names)
    sc_cols  = [c for c in df.columns if c not in injected]
    all_cols = sc_cols + WORKING_COLS + voucher_names

    # Styles
    header_font         = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    sc_header_fill      = PatternFill("solid", start_color="2D4A6B")   # Dark blue
    working_header_fill = PatternFill("solid", start_color="1F6B3E")   # Dark green
    vc_header_fill      = PatternFill("solid", start_color="7B2D8B")   # Purple
    eligible_fill       = PatternFill("solid", start_color="C6EFCE")   # Light green
    eligible_font       = Font(name="Arial", bold=True, color="276221", size=10)
    alt_fill            = PatternFill("solid", start_color="F2F2F2")   # Light grey alternate
    cell_font           = Font(name="Arial", size=10)
    center_align        = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align          = Alignment(horizontal="left", vertical="center")
    thin                = Side(border_style="thin", color="D0D0D0")
    border              = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Headers
    for ci, col in enumerate(all_cols, start=1):
        cell           = ws.cell(row=1, column=ci, value=col)
        cell.font      = header_font
        cell.alignment = center_align
        if col in sc_cols:
            cell.fill = sc_header_fill
        elif col in WORKING_COLS:
            cell.fill = working_header_fill
        else:
            cell.fill = vc_header_fill

    # Data rows
    for ri, (_, row) in enumerate(df[all_cols].iterrows(), start=2):
        is_alt = (ri % 2 == 0)
        for ci, col in enumerate(all_cols, start=1):
            val = row[col]
            if pd.isna(val):
                val = ""
            cell        = ws.cell(row=ri, column=ci, value=val)
            cell.border = border

            if col in voucher_names and str(val).strip().lower() == "yes-eligible":
                cell.fill      = eligible_fill
                cell.font      = eligible_font
                cell.alignment = center_align
            else:
                cell.font      = cell_font
                cell.alignment = left_align if col in sc_cols else center_align
                if is_alt and col not in voucher_names:
                    cell.fill = alt_fill

    # Auto-width (capped at 45)
    for ci, col in enumerate(all_cols, start=1):
        max_len = len(str(col))
        for ri in range(2, min(ws.max_row + 1, 200)):
            v = ws.cell(row=ri, column=ci).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 45)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_shopee_excel_output(df, voucher_configs, pid_flag_cols, dynamic_cols):
    """
    Return bytes of a styled Excel workbook for the Shopee (PID-level) pipeline.

    Column order: [SC Report columns] -> WORKING_COLS -> Product ID ->
    [4 mandatory PID flag columns] -> [dynamic exclusion _PID columns] -> [voucher columns]
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Shopee VC Eligibility"

    voucher_names = [vc["name"] for vc in voucher_configs if vc.get("name", "").strip()]

    shopee_working_cols = WORKING_COLS + ["Product ID"]
    injected = set(shopee_working_cols + pid_flag_cols + dynamic_cols + voucher_names)
    sc_cols  = [c for c in df.columns if c not in injected]
    all_cols = sc_cols + shopee_working_cols + pid_flag_cols + dynamic_cols + voucher_names

    # Styles
    header_font          = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    sc_header_fill        = PatternFill("solid", start_color="2D4A6B")   # Dark blue
    working_header_fill   = PatternFill("solid", start_color="1F6B3E")   # Dark green
    pid_flag_header_fill  = PatternFill("solid", start_color="C0392B")   # Red — mandatory exclusions
    dynamic_header_fill   = PatternFill("solid", start_color="B8860B")   # Amber — dynamic exclusions
    vc_header_fill        = PatternFill("solid", start_color="7B2D8B")   # Purple
    eligible_fill         = PatternFill("solid", start_color="C6EFCE")   # Light green
    eligible_font         = Font(name="Arial", bold=True, color="276221", size=10)
    flagged_fill          = PatternFill("solid", start_color="FCE4E4")   # Light red
    alt_fill              = PatternFill("solid", start_color="F2F2F2")   # Light grey alternate
    cell_font             = Font(name="Arial", size=10)
    center_align          = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align            = Alignment(horizontal="left", vertical="center")
    thin                  = Side(border_style="thin", color="D0D0D0")
    border                = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _header_fill_for(col):
        if col in sc_cols:
            return sc_header_fill
        if col in shopee_working_cols:
            return working_header_fill
        if col in pid_flag_cols:
            return pid_flag_header_fill
        if col in dynamic_cols:
            return dynamic_header_fill
        return vc_header_fill

    # Headers
    for ci, col in enumerate(all_cols, start=1):
        cell           = ws.cell(row=1, column=ci, value=col)
        cell.font      = header_font
        cell.alignment = center_align
        cell.fill      = _header_fill_for(col)

    flag_cols_set = set(pid_flag_cols + dynamic_cols)

    # Data rows
    for ri, (_, row) in enumerate(df[all_cols].iterrows(), start=2):
        is_alt = (ri % 2 == 0)
        for ci, col in enumerate(all_cols, start=1):
            val = row[col]
            if pd.isna(val):
                val = ""
            cell        = ws.cell(row=ri, column=ci, value=val)
            cell.border = border

            if col in voucher_names and str(val).strip().lower() == "yes-eligible":
                cell.fill      = eligible_fill
                cell.font      = eligible_font
                cell.alignment = center_align
            elif col in flag_cols_set and str(val).strip() != "":
                cell.fill      = flagged_fill
                cell.font      = cell_font
                cell.alignment = center_align
            else:
                cell.font      = cell_font
                cell.alignment = left_align if col in sc_cols else center_align
                if is_alt and col not in voucher_names:
                    cell.fill = alt_fill

    # Auto-width (capped at 40)
    for ci, col in enumerate(all_cols, start=1):
        max_len = len(str(col))
        for ri in range(2, min(ws.max_row + 1, 200)):
            v = ws.cell(row=ri, column=ci).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# Export — Eligible List (per voucher)
# ─────────────────────────────────────────────

_INVALID_FILENAME_CHARS_RE = re.compile(r'[\\/:*?"<>|]')


def _safe_filename_part(s):
    return _INVALID_FILENAME_CHARS_RE.sub("_", str(s)).strip()


def generate_eligible_export(df, marketplace, region, voucher_configs):
    """
    Build the "Export - Eligible List" output: one .xlsx per voucher, containing
    only the eligible SKUs (Shop SKU for Lazada, SellerSku for Zalora).

    File naming: <Marketplace>_<Region>_<Marketplace>_<Voucher Name>.xlsx
    The worksheet tab is renamed to exactly match the Voucher Name (Excel's
    31-character sheet-name limit is respected via truncation if needed).

    Returns (filename, file_bytes, is_zip):
      - If exactly one voucher yields output, returns a single .xlsx directly.
      - If multiple vouchers yield output, all files are bundled into a .zip.
    """
    import zipfile
    from openpyxl import Workbook

    if marketplace == "Shopee":
        sku_col = "Product ID" if "Product ID" in df.columns else find_product_id_col(df)
        export_header = "Product ID"
    else:
        sku_col = find_col(df, ["SellerSKU", "Seller SKU", "Shop SKU", "SKU", "seller_sku"])
        export_header = "Shop SKU" if marketplace == "Lazada" else "SellerSku"

    if not sku_col:
        raise ValueError(
            "Cannot find a SellerSKU / Shop SKU / Product ID column in the processed data to export."
        )

    files = {}
    for vc in voucher_configs:
        vc_name = vc.get("name", "").strip()
        if not vc_name or vc_name not in df.columns:
            continue

        eligible_mask = df[vc_name].astype(str).str.strip().str.lower() == "yes-eligible"
        eligible_skus = (
            df.loc[eligible_mask, sku_col]
            .astype(str)
            .str.strip()
        )
        eligible_skus = eligible_skus[~eligible_skus.isin(["", "nan", "NaN", "None"])]
        if marketplace == "Shopee":
            eligible_skus = eligible_skus.drop_duplicates()

        wb = Workbook()
        ws = wb.active
        ws.title = (vc_name[:31] if len(vc_name) > 31 else vc_name) or "Sheet1"
        ws.cell(row=1, column=1, value=export_header)
        for ri, sku in enumerate(eligible_skus.tolist(), start=2):
            ws.cell(row=ri, column=1, value=sku)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fname = _safe_filename_part(f"{marketplace}_{region}_{marketplace}_{vc_name}") + ".xlsx"
        files[fname] = buf.getvalue()

    if not files:
        raise ValueError("No eligible records found to export for any configured voucher.")

    if len(files) == 1:
        fname, data = next(iter(files.items()))
        return fname, data, False

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, data in files.items():
            zf.writestr(fname, data)
    zip_buf.seek(0)
    zip_name = _safe_filename_part(f"{marketplace}_{region}_Eligible_Export") + ".zip"
    return zip_name, zip_buf.getvalue(), True
